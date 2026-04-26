import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from tkinter import *
from tkcalendar import DateEntry
from datetime import date, datetime
from num2words import num2words
import pandas as pd
import openpyxl
import os
import tempfile
import subprocess
import sys
import matplotlib
matplotlib.use("TkAgg")
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4

# ─────────────────────────────────────────────
#  Constants
# ─────────────────────────────────────────────
DATA_FILE   = "form_data.xlsx"
SHEET_NAME  = "RESPONSES"
BG_COLOR    = "#2D9290"
SPLASH_BG   = "#16a085"

FUND_FIELDS = [
    "Sthanak Fund",
    "Jain Shala Fund",
    "Ayambil Fund",
    "Jeevdaya Fund",
    "Guru Vyavachee Fund",
    "Sadharan Fund",
    "Sabhasad Fee Savant",
]

# ─────────────────────────────────────────────
#  Root window (single, never destroyed)
# ─────────────────────────────────────────────
root = tk.Tk()
root.title("Billing System")
root.resizable(True, True)

# ─────────────────────────────────────────────
#  Helpers
# ─────────────────────────────────────────────
def clear_root():
    """Destroy all child widgets so a screen can redraw itself."""
    for w in root.winfo_children():
        w.destroy()


def load_customer_data():
    """Return {name: phone_str} from DATA_FILE, or {} on error."""
    data = {}
    if not os.path.exists(DATA_FILE):
        return data
    try:
        wb = openpyxl.load_workbook(DATA_FILE)
        if SHEET_NAME not in wb.sheetnames:
            return data
        for row in wb[SHEET_NAME].iter_rows(min_row=2, values_only=True):
            name, phone = row[0], row[1]
            if isinstance(name, str) and name.strip():
                data[name.strip()] = str(phone) if phone else ""
    except Exception as e:
        print("load_customer_data:", e)
    return data


def save_to_excel(form_data: dict):
    """Upsert a row in DATA_FILE by BillNo."""
    try:
        if os.path.exists(DATA_FILE):
            df = pd.read_excel(DATA_FILE, sheet_name=SHEET_NAME, dtype=str)
            # cast numeric columns back
            for col in FUND_FIELDS + ["BillNo", "Total Cost"]:
                if col in df.columns:
                    df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0).astype(int)

            mask = df["BillNo"] == int(form_data["BillNo"])
            if mask.any():
                for k, v in form_data.items():
                    df.loc[mask, k] = v
            else:
                df = pd.concat([df, pd.DataFrame([form_data])], ignore_index=True)
        else:
            df = pd.DataFrame([form_data])

        with pd.ExcelWriter(DATA_FILE, engine="openpyxl", mode="w") as writer:
            df.to_excel(writer, index=False, sheet_name=SHEET_NAME)

        messagebox.showinfo("Success", "Data saved successfully!")
    except Exception as e:
        messagebox.showerror("Error Saving", str(e))


# ─────────────────────────────────────────────
#  SCREEN 1 – Splash / Home
# ─────────────────────────────────────────────
def show_splash_screen():
    clear_root()
    root.geometry("600x400")
    root.configure(bg=SPLASH_BG)
    root.title("Billing System")

    def hover_in(e):  e.widget.configure(bg="#34495e")
    def hover_out(e): e.widget.configure(bg="#2c3e50")

    main = tk.Frame(root, bg=SPLASH_BG)
    main.pack(expand=True, fill="both", padx=20, pady=20)

    tk.Label(main, text="Billing System",
             font=("Times New Roman", 28, "bold"), fg="white", bg=SPLASH_BG).pack(pady=(20, 10))
    tk.Label(main, text="Professional Invoice Management Platform",
             font=("Times New Roman", 12), fg="#ecf0f1", bg=SPLASH_BG).pack(pady=(0, 40))

    btn_style = dict(font=("Times New Roman", 16, "bold"),
                     padx=20, pady=15, width=15,
                     bg="#2c3e50", fg="white", relief="flat", cursor="hand2")

    btn_frame = tk.Frame(main, bg=SPLASH_BG)
    btn_frame.pack(expand=True)

    b1 = tk.Button(btn_frame, text="Make Receipt", command=show_receipt_ui, **btn_style)
    b1.grid(row=0, column=0, padx=20, pady=20)

    b2 = tk.Button(btn_frame, text="Analysis", command=show_reports_ui, **btn_style)
    b2.grid(row=0, column=1, padx=20, pady=20)

    for b in (b1, b2):
        b.bind("<Enter>", hover_in)
        b.bind("<Leave>", hover_out)

    tk.Label(main, text="© 2024 Billing System",
             font=("Times New Roman", 10), fg="#bdc3c7", bg=SPLASH_BG).pack(side="bottom", pady=10)


# ─────────────────────────────────────────────
#  SCREEN 2 – Receipt UI
# ─────────────────────────────────────────────
def show_receipt_ui(edit_data=None):
    clear_root()
    root.geometry("1200x750")
    root.configure(bg=BG_COLOR)
    root.title("Billing System – Receipt")

    # ── receipt_lines is LOCAL to this screen call ─────────
    # Using nonlocal in inner functions lets generate_receipt,
    # print_receipt, download_pdf, and reset_form all share
    # the same list without touching any module-level variable.
    receipt_lines = []

    # ── Variables ──────────────────────────────────
    fund_vars   = {f: IntVar(value=0) for f in FUND_FIELDS}
    user_name   = StringVar()
    phone_number= StringVar()
    total_cost  = StringVar(value="0")
    bill_no     = IntVar(value=0)
    date_of_bill= StringVar(value=datetime.now().strftime("%d-%m-%Y"))

    def enforce_upper(*_):
        user_name.set(user_name.get().upper())
    user_name.trace_add("write", enforce_upper)

    def enforce_phone(*_):
        val = phone_number.get()
        # Keep only digits, truncate to 10
        cleaned = "".join(c for c in val if c.isdigit())[:10]
        if val != cleaned:
            phone_number.set(cleaned)
    phone_number.trace_add("write", enforce_phone)

    if edit_data:
        bill_no.set(int(edit_data.get("BillNo", 0)))
        user_name.set(str(edit_data.get("Name", "")))
        phone_number.set(str(edit_data.get("Phone", "")))
        date_of_bill.set(str(edit_data.get("Bill Date", datetime.now().strftime("%d-%m-%Y"))))
        for f in FUND_FIELDS:
            fund_vars[f].set(int(edit_data.get(f, 0)))
        total_cost.set(str(sum(fund_vars[f].get() for f in FUND_FIELDS)))

    # ── Customer autocomplete ───────────────────────
    customer_data = load_customer_data()

    def update_suggestions(event=None):
        typed = user_name.get().strip().lower()
        suggestion_box.delete(0, END)
        if not typed:
            suggestion_box.place_forget(); return
        matches = [n for n in customer_data if typed in n.lower()]
        if matches:
            for n in matches:
                suggestion_box.insert(END, n)
            sx = name_entry.winfo_rootx() - root.winfo_rootx()
            sy = name_entry.winfo_rooty() - root.winfo_rooty() + name_entry.winfo_height()
            suggestion_box.place(x=sx, y=sy)
            suggestion_box.tkraise()
        else:
            suggestion_box.place_forget()

    def on_suggestion_select(event=None):
        if not suggestion_box.curselection(): return
        sel = suggestion_box.get(suggestion_box.curselection()[0])
        user_name.set(sel)
        phone_number.set(customer_data.get(sel, ""))
        suggestion_box.place_forget()

    def hide_suggestions(event=None):
        if event and event.widget not in (name_entry, suggestion_box):
            suggestion_box.place_forget()

    root.bind("<Button-1>", hide_suggestions)

    # ── Validation helpers ──────────────────────────
    def validate_name(n):   return n.replace(" ", "").isalpha()
    def validate_phone(p):  return p.isdigit() and len(p) == 10
    def validate_total(t):  return str(t).isdigit() and int(t) > 0

    # ── Receipt logic ───────────────────────────────
    def display_receipt(lines):
        for w in receipt_frame.winfo_children():
            w.destroy()
        for i, line in enumerate(lines):
            Label(receipt_frame, text=line, font=("Courier", 11),
                  anchor="w", bg="white").grid(row=i, column=0, sticky="w", padx=10)

    def generate_receipt():
        nonlocal receipt_lines

        if all(fund_vars[f].get() == 0 for f in FUND_FIELDS):
            messagebox.showerror("Error", "Please enter at least one fund amount.")
            return

        t = sum(fund_vars[f].get() for f in FUND_FIELDS)
        total_txt.config(state="normal")
        total_cost.set(str(t))
        total_txt.config(state="readonly")

        name  = user_name.get().strip().upper()
        phone = phone_number.get().strip()
        total = total_cost.get()
        bno   = bill_no.get()

        if bno == 0:
            messagebox.showerror("Invalid Input", "Please enter a valid Bill Number."); return
        if not validate_phone(phone):
            messagebox.showerror("Invalid Input", "Phone number must be exactly 10 digits."); return
        if not validate_name(name):
            messagebox.showerror("Invalid Input", "Name must contain only alphabets."); return
        if not validate_total(total):
            messagebox.showerror("Invalid Input", "Total amount must be greater than zero."); return

        receipt_lines = []
        receipt_lines.append("|| Om Arham ||".center(60))
        receipt_lines.append("Shree Shwetambar Sthanakwasi Jain Sangh".center(60))
        receipt_lines.append("544 E, Vyapari Peth, Shahupuri, Kolhapur".center(60))
        receipt_lines.append("-" * 60)
        receipt_lines.append(f"{'Receipt No':<15}: {bno}")
        receipt_lines.append(f"{'Date':<15}: {date_of_bill.get()}")
        receipt_lines.append(f"{'Name':<15}: {name}")
        receipt_lines.append(f"{'Phone No':<15}: {phone}")
        receipt_lines.append("-" * 60)
        receipt_lines.append(f"{'Type of Fund':<40}{'Amount':>10}")
        receipt_lines.append("-" * 60)

        for f in FUND_FIELDS:
            if fund_vars[f].get() > 0:
                receipt_lines.append(f"{f:<40}{fund_vars[f].get():>10}")

        receipt_lines.append("-" * 60)
        receipt_lines.append(f"{'Total Amount':<40}{total:>10}")
        receipt_lines.append("-" * 60)
        total_words = num2words(int(total), lang="en").capitalize()
        receipt_lines.append(f"Total in Words: {total_words} only")
        receipt_lines.append("-" * 60)

        display_receipt(receipt_lines)

        form_data = {
            "Name": name, "Phone": phone, "BillNo": bno,
            "Bill Date": date_of_bill.get(),
            **{f: fund_vars[f].get() for f in FUND_FIELDS},
            "Total Cost": int(total)
        }
        save_to_excel(form_data)
        # refresh autocomplete with newly saved name
        customer_data.update(load_customer_data())

    def print_receipt():
        if not receipt_lines:
            messagebox.showerror("Error", "Please generate a receipt first."); return
        content  = "\n".join(receipt_lines)
        filename = tempfile.mktemp(".txt")
        with open(filename, "w") as f:
            f.write(content)
        os.startfile(filename, "print")
        reset_form()

    def download_pdf():
        if not receipt_lines:
            messagebox.showerror("Error", "Please generate a receipt first."); return
        file_path = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            initialfile=f"Receipt_{bill_no.get()}.pdf",
            filetypes=[("PDF Files", "*.pdf")],
            title="Save Receipt As PDF"
        )
        if not file_path:
            return  # user cancelled
        c = canvas.Canvas(file_path, pagesize=A4)
        obj = c.beginText(40, 800)
        obj.setFont("Courier", 12)
        for line in receipt_lines:
            obj.textLine(line)
        c.drawText(obj)
        c.showPage()
        c.save()
        messagebox.showinfo("PDF Saved", f"Receipt saved as:\n{file_path}")
        reset_form()

    def reset_form():
        nonlocal receipt_lines
        receipt_lines = []
        for w in receipt_frame.winfo_children():
            w.destroy()
        for f in FUND_FIELDS:
            fund_vars[f].set(0)
        total_cost.set("0")
        user_name.set("")
        phone_number.set("")
        bill_no.set(0)
        date_of_bill.set(datetime.now().strftime("%d-%m-%Y"))

    # ── Layout ─────────────────────────────────────
    # Header
    header = Frame(root, bg=BG_COLOR, bd=12, relief=GROOVE)
    header.pack(fill=X, pady=2)
    Button(header, text="< Back", font=("Times New Roman", 16, "bold"),
           bg="yellow", fg="crimson", padx=5, pady=3, width=14,
           command=show_splash_screen).pack(side=LEFT, padx=10)
    Label(header, text="Billing System", bg=BG_COLOR, fg="white",
          font=("Times New Roman", 30, "bold")).pack(side=LEFT, padx=10, expand=True)

    # Customer details
    f1 = LabelFrame(root, text="Customer Details",
                    font=("Times New Roman", 16, "bold"), bd=10, bg=BG_COLOR, fg="gold")
    f1.place(x=0, y=80, relwidth=1)

    Label(f1, text="Customer Name", font=("Times New Roman", 16),
          bg=BG_COLOR, fg="white").grid(row=0, column=0, padx=10, pady=5)
    name_entry = Entry(f1, textvariable=user_name, font=("Times New Roman", 16), width=16)
    name_entry.grid(row=0, column=1, padx=10, pady=5)
    suggestion_box = Listbox(root, font=("Times New Roman", 14), width=23, height=5)
    suggestion_box.place_forget()
    name_entry.bind("<KeyRelease>", update_suggestions)
    suggestion_box.bind("<<ListboxSelect>>", on_suggestion_select)

    Label(f1, text="Phone No.", font=("Times New Roman", 16),
          bg=BG_COLOR, fg="white").grid(row=0, column=2, padx=10, pady=5)
    Entry(f1, width=13, font=("Times New Roman", 16), relief=SUNKEN, bd=7,
          textvariable=phone_number).grid(row=0, column=3, padx=10, pady=5)

    Label(f1, text="Receipt No.", font=("Times New Roman", 16),
          bg=BG_COLOR, fg="white").grid(row=0, column=4, padx=10, pady=5)
    Entry(f1, width=8, font=("Times New Roman", 16), relief=SUNKEN, bd=7,
          textvariable=bill_no).grid(row=0, column=5, padx=10, pady=5)

    Label(f1, text="Date", font=("Times New Roman", 16),
          bg=BG_COLOR, fg="white").grid(row=0, column=6, padx=10, pady=5)
    DateEntry(f1, width=13, font=("Times New Roman", 16), relief=SUNKEN, bd=7,
              textvariable=date_of_bill, date_pattern="dd-mm-yyyy",
              maxdate=date.today(), background="darkblue",
              foreground="white").grid(row=0, column=7, padx=10, pady=5)

    # Fund entry panel
    f2 = LabelFrame(root, text="Product Details",
                    font=("Times New Roman", 16, "bold"), fg="gold", bg=BG_COLOR, bd=10)
    f2.place(x=0, y=175, width=600, height=490)

    Label(f2, text="Category", font=("Times New Roman", 16, "underline", "bold"),
          fg="black", bg=BG_COLOR).grid(row=0, column=0, padx=20, pady=5)
    Label(f2, text="Amount", font=("Times New Roman", 16, "underline", "bold"),
          fg="black", bg=BG_COLOR).grid(row=0, column=1, padx=200, pady=5)

    for i, f in enumerate(FUND_FIELDS, start=1):
        Label(f2, text=f, font=("Times New Roman", 16),
              fg="black", bg=BG_COLOR).grid(row=i, column=0, padx=20, pady=5)
        Entry(f2, width=16, font=("Times New Roman", 16), relief=SUNKEN, bd=7,
              justify=CENTER, textvariable=fund_vars[f]).grid(row=i, column=1, padx=20, pady=5)

    Label(f2, text="Total Amount", font=("Times New Roman", 16),
          fg="black", bg=BG_COLOR).grid(row=len(FUND_FIELDS)+1, column=0, padx=20, pady=5)
    total_txt = Entry(f2, width=16, font=("Times New Roman", 16), relief=SUNKEN, bd=7,
                      justify=CENTER, textvariable=total_cost, state="readonly")
    total_txt.grid(row=len(FUND_FIELDS)+1, column=1, padx=20, pady=5)

    # Scrollable receipt panel
    receipt_container = Frame(root, bd=10, relief=GROOVE)
    receipt_container.place(x=620, y=175, width=580, height=490)
    receipt_canvas  = Canvas(receipt_container, bg="white")
    receipt_canvas.pack(side=LEFT, fill=BOTH, expand=True)
    receipt_scrollbar = Scrollbar(receipt_container, orient=VERTICAL, command=receipt_canvas.yview)
    receipt_scrollbar.pack(side=RIGHT, fill=Y)
    receipt_canvas.configure(yscrollcommand=receipt_scrollbar.set)
    receipt_canvas.bind("<Configure>",
                        lambda e: receipt_canvas.configure(scrollregion=receipt_canvas.bbox("all")))
    receipt_frame = Frame(receipt_canvas, bg="white")
    receipt_canvas.create_window((0, 0), window=receipt_frame, anchor="nw")

    def _on_mousewheel(event):
        receipt_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
    receipt_canvas.bind_all("<MouseWheel>", _on_mousewheel)
    receipt_canvas.bind_all("<Button-4>", lambda e: receipt_canvas.yview_scroll(-1, "units"))
    receipt_canvas.bind_all("<Button-5>", lambda e: receipt_canvas.yview_scroll(1, "units"))

    # Bottom action bar
    f4 = Frame(root, relief=GROOVE, bd=10, bg=BG_COLOR)
    f4.pack(side=BOTTOM, fill=X, pady=10)
    for i in range(4):
        f4.grid_columnconfigure(i, weight=1, uniform="equal")

    btn_cfg = dict(font=("Times New Roman", 16, "bold"),
                   bg="yellow", fg="crimson", padx=7, pady=5, width=14)
    Button(f4, text="Receipt",      command=generate_receipt, **btn_cfg).grid(row=0, column=0, padx=10, pady=10, sticky="ew")
    Button(f4, text="Print",        command=print_receipt,    **btn_cfg).grid(row=0, column=1, padx=10, pady=10, sticky="ew")
    Button(f4, text="Reset",        command=reset_form,       **btn_cfg).grid(row=0, column=2, padx=10, pady=10, sticky="ew")
    Button(f4, text="Download PDF", command=download_pdf,     **btn_cfg).grid(row=0, column=3, padx=10, pady=10, sticky="ew")


# ─────────────────────────────────────────────
#  SCREEN 3 – Reports / Analysis
# ─────────────────────────────────────────────
def show_reports_ui():
    clear_root()
    root.title("Billing System – Reports & Analysis")
    root.geometry("1300x750")
    root.configure(bg=SPLASH_BG)

    # Load fresh data
    try:
        df = pd.read_excel(DATA_FILE, sheet_name=SHEET_NAME)
        for f in FUND_FIELDS:
            if f in df.columns:
                df[f] = pd.to_numeric(df[f], errors="coerce").fillna(0).astype(int)
        df["Bill Date"] = df["Bill Date"].astype(str)
    except FileNotFoundError:
        messagebox.showwarning("No Data", f"'{DATA_FILE}' not found. No records yet.")
        show_splash_screen()
        return
    except Exception as e:
        messagebox.showerror("Load Error", str(e))
        show_splash_screen()
        return

    df["Edit"]   = "✎ Edit"
    df["Delete"] = "🗑 Delete"

    # ── Style ──────────────────────────────────────
    style = ttk.Style()
    style.theme_use("clam")
    style.configure("Treeview", background="#ecf0f1", foreground="#2c3e50",
                    font=("Times New Roman", 10))
    style.configure("Treeview.Heading", background="#34495e", foreground="white",
                    font=("Times New Roman", 11, "bold"))

    # ── Main layout ─────────────────────────────────
    main_frame = tk.Frame(root, bg=SPLASH_BG, padx=10, pady=10)
    main_frame.pack(fill="both", expand=True)

    # Header
    header_frame = tk.Frame(main_frame, bg=SPLASH_BG)
    header_frame.pack(fill="x", pady=(0, 10))

    tk.Button(header_frame, text="← Back",
              font=("Times New Roman", 12, "bold"),
              padx=15, pady=5, bg="#e74c3c", fg="white",
              relief="flat", cursor="hand2",
              command=show_splash_screen           # ← FIX: same-process navigation
              ).pack(side="left")

    tk.Label(header_frame, text="Reports & Analysis",
             font=("Times New Roman", 20, "bold"),
             fg="white", bg=SPLASH_BG).pack(expand=True)

    # ── Search / filter bar ─────────────────────────
    search_frame = tk.LabelFrame(main_frame, text="Search & Filter",
                                 font=("Times New Roman", 12, "bold"),
                                 bg=SPLASH_BG, fg="white", padx=10, pady=5)
    search_frame.pack(fill="x", pady=(0, 10))

    tk.Label(search_frame, text="Search:", font=("Times New Roman", 11),
             bg=SPLASH_BG, fg="white").grid(row=0, column=0, sticky="w", padx=(0, 5))
    e1 = tk.Entry(search_frame, width=15, font=("Times New Roman", 11), relief="flat")
    e1.grid(row=0, column=1, padx=(0, 15))

    tk.Label(search_frame, text="Start Date:", font=("Times New Roman", 11),
             bg=SPLASH_BG, fg="white").grid(row=0, column=2, sticky="w", padx=(0, 5))
    start_date_entry = tk.Entry(search_frame, width=12, font=("Times New Roman", 11), relief="flat")
    start_date_entry.grid(row=0, column=3, padx=(0, 15))

    tk.Label(search_frame, text="End Date:", font=("Times New Roman", 11),
             bg=SPLASH_BG, fg="white").grid(row=0, column=4, sticky="w", padx=(0, 5))
    end_date_entry = tk.Entry(search_frame, width=12, font=("Times New Roman", 11), relief="flat")
    end_date_entry.grid(row=0, column=5, padx=(0, 15))

    tk.Label(search_frame, text="Filter by Fund:", font=("Times New Roman", 11),
             bg=SPLASH_BG, fg="white").grid(row=0, column=6, sticky="w", padx=(0, 5))
    fund_filter = ttk.Combobox(search_frame, width=18, font=("Times New Roman", 11),
                               values=["All"] + FUND_FIELDS, state="readonly")
    fund_filter.set("All")
    fund_filter.grid(row=0, column=7, padx=(0, 15))

    tk.Button(search_frame, text="Search",
              font=("Times New Roman", 11, "bold"), padx=20, pady=5,
              bg="#3498db", fg="white", relief="flat", cursor="hand2",
              command=lambda: do_search()).grid(row=0, column=8, padx=10)

    tk.Button(search_frame, text="⬇ Download Results",
              font=("Times New Roman", 11, "bold"), padx=10, pady=5,
              bg="#27ae60", fg="white", relief="flat", cursor="hand2",
              command=lambda: download_results()).grid(row=0, column=9, padx=10)

    # ── Results treeview ────────────────────────────
    results_frame = tk.Frame(main_frame, bg=SPLASH_BG)
    results_frame.pack(fill="x", pady=0)

    tree_frame = tk.Frame(results_frame, bg=SPLASH_BG)
    tree_frame.pack(fill="x", expand=False)
    tree_frame.grid_rowconfigure(0, weight=0)
    tree_frame.grid_columnconfigure(0, weight=1)

    trv = ttk.Treeview(tree_frame, selectmode="browse")
    trv.grid(row=0, column=0, sticky="ew")

    v_scroll = ttk.Scrollbar(tree_frame, orient="vertical",   command=trv.yview)
    v_scroll.grid(row=0, column=1, sticky="ns")
    h_scroll = ttk.Scrollbar(tree_frame, orient="horizontal", command=trv.xview)
    h_scroll.grid(row=1, column=0, sticky="ew")
    trv.configure(yscrollcommand=v_scroll.set, xscrollcommand=h_scroll.set)

    trv["columns"] = list(df.columns)
    trv["show"]    = "headings"
    for col in df.columns:
        trv.column(col, width=100, anchor="center")
        trv.heading(col, text=col)

    # ── Individual summary treeview ──────────────────
    summary_frame = tk.LabelFrame(main_frame, text="Individual Donation Summary",
                                  font=("Times New Roman", 12, "bold"),
                                  bg=SPLASH_BG, fg="white")
    summary_frame.pack(fill="x", expand=False, pady=(10, 0))

    sum_container = tk.Frame(summary_frame, bg=SPLASH_BG)
    sum_container.pack(fill="both", expand=True)

    summary_tree = ttk.Treeview(sum_container, selectmode="browse", height=1)
    summary_tree.pack(side="left", fill="both", expand=True)
    sum_scroll = ttk.Scrollbar(sum_container, orient="vertical", command=summary_tree.yview)
    sum_scroll.pack(side="right", fill="y")
    summary_tree.configure(yscrollcommand=sum_scroll.set)

    sum_cols = ["Name"] + FUND_FIELDS + ["Total Donated"]
    summary_tree["columns"] = sum_cols
    summary_tree["show"]    = "headings"
    for col in sum_cols:
        summary_tree.heading(col, text=col)
        summary_tree.column(col, anchor="center", width=100)

    # ── Fund totals treeview ─────────────────────────
    fund_frame = tk.LabelFrame(main_frame, text="Total Donations by Fund",
                               font=("Times New Roman", 12, "bold"),
                               bg=SPLASH_BG, fg="white")
    fund_frame.pack(fill="x", expand=False, pady=(10, 20))

    fund_tree = ttk.Treeview(fund_frame, columns=("Fund", "Total"), show="headings", height=7)
    fund_tree.pack(fill="x", expand=True)
    fund_tree.heading("Fund",  text="Fund Name")
    fund_tree.heading("Total", text="Total Amount")
    fund_tree.column("Fund",  anchor="center", width=300)
    fund_tree.column("Total", anchor="center", width=200)

    # ── Inner functions ──────────────────────────────
    def populate_trv(source_df):
        for item in trv.get_children():
            trv.delete(item)
        for row in source_df.to_numpy().tolist():
            trv.insert("", "end", values=row)
        trv.config(height=min(max(len(source_df), 1), 10))

    def do_search():
        query      = e1.get().strip()
        start_date = start_date_entry.get().strip()
        end_date   = end_date_entry.get().strip()
        fund_sel   = fund_filter.get()

        filt = df.copy()

        if fund_sel != "All" and fund_sel in filt.columns:
            filt = filt[filt[fund_sel] > 0]

        if start_date and end_date:
            try:
                filt = filt[(filt["Bill Date"] >= start_date) &
                            (filt["Bill Date"] <= end_date)]
            except KeyError:
                messagebox.showerror("Error", "Column 'Bill Date' not found!"); return

        if query:
            if query.isdigit():
                filt = filt[filt["BillNo"] == int(query)]
            else:
                filt = filt[filt["Name"].str.contains(query, case=False, na=False)]

        populate_trv(filt)

    def download_results():
        data = [trv.item(c)["values"] for c in trv.get_children()]
        if not data:
            messagebox.showinfo("No Data", "No results to export."); return
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            title="Save Search Results As")
        if not file_path: return
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Search Results"
        ws.append(list(trv["columns"]))
        for row in data:
            ws.append(row)
        wb.save(file_path)
        messagebox.showinfo("Success", f"Saved to:\n{file_path}")

    def on_treeview_click(event):
        item = trv.identify_row(event.y)
        col  = trv.identify_column(event.x)
        if not item or not col: return
        col_idx  = int(col.replace("#", "")) - 1
        values   = trv.item(item)["values"]
        if not values: return
        col_name = trv["columns"][col_idx]

        if col_name == "Edit":
            edit_data = dict(zip(trv["columns"], values))
            show_receipt_ui(edit_data=edit_data)

        elif col_name == "Delete":
            bno = values[list(trv["columns"]).index("BillNo")]
            if messagebox.askyesno("Confirm Delete", f"Delete Bill No {bno}?"):
                delete_bill(bno)

    def delete_bill(bill_number):
        try:
            wb = openpyxl.load_workbook(DATA_FILE)
            if SHEET_NAME not in wb.sheetnames:
                messagebox.showerror("Error", f"Sheet '{SHEET_NAME}' not found."); return
            ws  = wb[SHEET_NAME]
            row_to_delete = None
            for row in range(2, ws.max_row + 1):
                if str(ws.cell(row=row, column=3).value) == str(bill_number):
                    row_to_delete = row; break
            if row_to_delete:
                ws.delete_rows(row_to_delete)
                ws.title = SHEET_NAME
                wb.save(DATA_FILE)
                messagebox.showinfo("Deleted", f"Bill No {bill_number} deleted.")
                show_reports_ui()   # reload screen with fresh data
            else:
                messagebox.showwarning("Not Found", f"Bill No {bill_number} not found.")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def load_summary():
        if "Name" not in df.columns: return
        grouped = df.groupby("Name")[FUND_FIELDS].sum()
        grouped["Total Donated"] = grouped.sum(axis=1)
        grouped = grouped.sort_values("Total Donated", ascending=False)
        for item in summary_tree.get_children():
            summary_tree.delete(item)
        for name, row in grouped.iterrows():
            vals = [name] + [int(row[f]) for f in FUND_FIELDS] + [int(row["Total Donated"])]
            summary_tree.insert("", "end", values=vals)
        summary_tree.config(height=min(len(grouped), 10))

    def load_fund_totals():
        for item in fund_tree.get_children():
            fund_tree.delete(item)
        for fund, amount in df[FUND_FIELDS].sum().items():
            fund_tree.insert("", "end", values=(fund, int(amount)))

    trv.bind("<ButtonRelease-1>", on_treeview_click)
    fund_filter.bind("<<ComboboxSelected>>", lambda e: do_search())

    # Initial population
    do_search()
    load_summary()
    load_fund_totals()


# ─────────────────────────────────────────────
#  Entry point
# ─────────────────────────────────────────────
if __name__ == "__main__":
    show_splash_screen()
    root.mainloop()